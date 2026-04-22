#!/usr/bin/env python3
"""
test_dashboard.py — Garmin Local Archive — Dashboard Pipeline Test

Run from the project folder:
    python tests/test_dashboard.py

No network, no GUI, no Garmin API calls.
Uses synthetic raw JSON files — no real data required.
Cleans up after itself — leaves no files behind.
"""

import json
import os
import sys
import shutil
import tempfile
import logging
from pathlib import Path

# ── Path setup ─────────────────────────────────────────────────────────────────
_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(_ROOT / "garmin"))
sys.path.insert(0, str(_ROOT))
sys.path.insert(0, str(_ROOT / "layouts"))
sys.path.insert(0, str(_ROOT / "dashboards"))
logging.disable(logging.CRITICAL)

# ── Results tracking ───────────────────────────────────────────────────────────
_pass = 0
_fail = 0
_failures = []

def check(name, condition):
    global _pass, _fail
    if condition:
        _pass += 1
        print(f"  ✓  {name}")
    else:
        _fail += 1
        _failures.append(name)
        print(f"  ✗  {name}")

def section(title):
    print(f"\n{'─' * 55}")
    print(f"  {title}")
    print(f"{'─' * 55}")

# ── Temp directory as BASE_DIR ─────────────────────────────────────────────────
_TMPDIR = Path(tempfile.mkdtemp(prefix="garmin_dash_test_"))
os.environ["GARMIN_OUTPUT_DIR"] = str(_TMPDIR)

import importlib
import garmin_config as cfg
importlib.reload(cfg)

# ── Synthetic raw data ─────────────────────────────────────────────────────────

_TEST_DATE = "2026-03-01"

_RAW = {
    "date": _TEST_DATE,
    "sleep": {
        "dailySleepDTO": {
            "sleepTimeSeconds":  27000,
            "deepSleepSeconds":  5400,
            "lightSleepSeconds": 13500,
            "remSleepSeconds":   6750,
            "awakeSleepSeconds": 1350,
        }
    },
    "heart_rates": {
        "heartRateValues": [
            [1740787200000, 58],
            [1740787260000, 60],
            [1740787320000, 62],
        ]
    },
    "stress": {
        "stressValuesArray": [
            [1740787200000, 25],
            [1740787260000, 30],
        ],
        "stressChartValueOffset": 0,
        "bodyBatteryValuesArray": [
            [1740787200000, "CHARGED", 85, 1],
            [1740787260000, "CHARGED", 83, 1],
        ],
    },
    "spo2": {
        "spO2HourlyAverages": [
            {"startGMT": "2026-03-01T08:00:00", "spO2Reading": 97},
            {"startGMT": "2026-03-01T09:00:00", "spO2Reading": 98},
        ]
    },
    "respiration": {
        "respirationValuesArray": [
            {"startGMT": "2026-03-01T08:00:00", "respirationValue": 14.5},
            {"startGMT": "2026-03-01T09:00:00", "respirationValue": 15.0},
        ]
    },
}

def _write_raw(base_dir: Path):
    raw_dir = base_dir / "garmin_data" / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    f = raw_dir / f"garmin_raw_{_TEST_DATE}.json"
    f.write_text(json.dumps(_RAW), encoding="utf-8")
    return raw_dir

_NULL_DATE = "2026-03-02"

_NULL_RAW = {
    "date": _NULL_DATE,
    "sleep": {
        "dailySleepDTO": {
            "sleepTimeSeconds":  27000,
            "deepSleepSeconds":  5400,
            "lightSleepSeconds": 13500,
            "remSleepSeconds":   6750,
            "awakeSleepSeconds": 1350,
        }
    },
    "heart_rates": {
        "heartRateValues": None,
    },
    "stress": {
        "stressValuesArray": None,
        "stressChartValueOffset": 0,
        "bodyBatteryValuesArray": None,
    },
    "spo2":        {},
    "respiration": {},
}

def _write_null_raw(base_dir: Path):
    raw_dir = base_dir / "garmin_data" / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    f = raw_dir / f"garmin_raw_{_NULL_DATE}.json"
    f.write_text(json.dumps(_NULL_RAW), encoding="utf-8")


# ══════════════════════════════════════════════════════════════════════════════
#  1. garmin_map — intraday normalization
# ══════════════════════════════════════════════════════════════════════════════

section("1. garmin_map — intraday normalization")

_write_raw(_TMPDIR)
importlib.reload(cfg)

from maps import garmin_map

result_hr = garmin_map.get("heart_rate_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
hr_series = result_hr["values"][0]["series"]

check("heart_rate: result has values",          isinstance(result_hr["values"], list))
check("heart_rate: series is list",             isinstance(hr_series, list))
check("heart_rate: series not empty",           len(hr_series) > 0)
check("heart_rate: entry has ts key",           "ts" in hr_series[0])
check("heart_rate: entry has value key",        "value" in hr_series[0])
check("heart_rate: ts is ISO string",           isinstance(hr_series[0]["ts"], str) and "T" in hr_series[0]["ts"])
check("heart_rate: value is float",             isinstance(hr_series[0]["value"], float))
check("heart_rate: value correct",              hr_series[0]["value"] == 58.0)

result_stress = garmin_map.get("stress_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
stress_series = result_stress["values"][0]["series"]
check("stress: series not empty",               len(stress_series) > 0)
check("stress: value is float",                 isinstance(stress_series[0]["value"], float))

result_bb = garmin_map.get("body_battery_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
bb_series = result_bb["values"][0]["series"]
check("body_battery: series not empty",         len(bb_series) > 0)
check("body_battery: val_index=2 correct",      bb_series[0]["value"] == 85.0)

result_spo2 = garmin_map.get("spo2_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
spo2_series = result_spo2["values"][0]["series"]
check("spo2: series not empty",                 len(spo2_series) > 0)
check("spo2: dict-based extraction works",      spo2_series[0]["value"] == 97.0)

result_resp = garmin_map.get("respiration_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
resp_series = result_resp["values"][0]["series"]
check("respiration: series not empty",          len(resp_series) > 0)
check("respiration: value correct",             resp_series[0]["value"] == 14.5)

result_missing = garmin_map.get("heart_rate_series", "2000-01-01", "2000-01-01", resolution="intraday")
check("missing date: series is None",           result_missing["values"][0]["series"] is None)

# null-Wert-Robustheit (F + K) — separates raw-File
_write_null_raw(_TMPDIR)

result_hr_null = garmin_map.get("heart_rate_series", _NULL_DATE, _NULL_DATE, resolution="intraday")
check("null heart_rate: series is None",        result_hr_null["values"][0]["series"] is None)

result_stress_null = garmin_map.get("stress_series", _NULL_DATE, _NULL_DATE, resolution="intraday")
check("null stress: series is None",            result_stress_null["values"][0]["series"] is None)

result_bb_null = garmin_map.get("body_battery_series", _NULL_DATE, _NULL_DATE, resolution="intraday")
check("null body_battery: series is None",      result_bb_null["values"][0]["series"] is None)

result_spo2_null = garmin_map.get("spo2_series", _NULL_DATE, _NULL_DATE, resolution="intraday")
check("null spo2: series is None",              result_spo2_null["values"][0]["series"] is None)

result_resp_null = garmin_map.get("respiration_series", _NULL_DATE, _NULL_DATE, resolution="intraday")
check("null respiration: series is None",       result_resp_null["values"][0]["series"] is None)


# ══════════════════════════════════════════════════════════════════════════════
#  2. field_map — routing
# ══════════════════════════════════════════════════════════════════════════════

section("2. field_map — routing to garmin_map")

from maps import field_map

result_fm = field_map.get("heart_rate_series", _TEST_DATE, _TEST_DATE, resolution="intraday")
check("field_map: garmin key in result",        "garmin" in result_fm)
check("field_map: values present",              len(result_fm["garmin"]["values"]) > 0)
check("field_map: fallback key present",        "fallback" in result_fm["garmin"])


# ══════════════════════════════════════════════════════════════════════════════
#  3. dash_layout — tokens and text
# ══════════════════════════════════════════════════════════════════════════════

section("3. dash_layout — design tokens")

sys.path.insert(0, str(_ROOT / "layouts"))
import dash_layout as layout

check("metric_meta heart_rate_series",          layout.get_metric_meta("heart_rate_series") != {})
check("metric_meta has label",                  "label" in layout.get_metric_meta("heart_rate_series"))
check("metric_meta has color",                  "color" in layout.get_metric_meta("heart_rate_series"))
check("metric_meta unknown returns {}",         layout.get_metric_meta("nonexistent") == {})
check("excel row color heart_rate",             layout.get_excel_row_color("heart_rate_series") != "")
check("excel row color unknown = white",        layout.get_excel_row_color("nonexistent") == "FFFFFF")
check("disclaimer is string",                   isinstance(layout.get_disclaimer(), str))
check("footer html contains anchor",            "<a " in layout.get_footer(html=True))
check("footer plain no anchor",                 "<a " not in layout.get_footer(html=False))


# ══════════════════════════════════════════════════════════════════════════════
#  4. dash_layout_html — CSS and builders
# ══════════════════════════════════════════════════════════════════════════════

section("4. dash_layout_html — HTML assets")

import dash_layout_html as layout_html

check("CSS is string",                          isinstance(layout_html.get_css(), str))
check("CSS contains body",                      "body" in layout_html.get_css())
check("plotly CDN is URL",                      layout_html.get_plotly_cdn().startswith("https://"))
header = layout_html.build_header("Test Title", "Test Subtitle")
check("header contains title",                  "Test Title" in header)
check("header contains subtitle",               "Test Subtitle" in header)
disclaimer = layout_html.build_disclaimer("Test disclaimer")
check("disclaimer contains text",               "Test disclaimer" in disclaimer)


# ══════════════════════════════════════════════════════════════════════════════
#  5. specialist — build()
# ══════════════════════════════════════════════════════════════════════════════

section("5. timeseries_garmin specialist — build()")

import importlib.util
_spec = importlib.util.spec_from_file_location(
    "timeseries_dash",
    _ROOT / "dashboards" / "timeseries_garmin_html-xls_dash.py"
)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

check("META present",                           hasattr(_mod, "META"))
check("META has name",                          "name" in _mod.META)
check("META has formats",                       "formats" in _mod.META)
check("META formats has html",                  "html" in _mod.META["formats"])
check("META formats has excel",                 "excel" in _mod.META["formats"])

data = _mod.build(_TEST_DATE, _TEST_DATE, {})
check("build returns dict",                     isinstance(data, dict))
check("build has title",                        "title" in data)
check("build has fields",                       "fields" in data)
check("build fields is list",                   isinstance(data["fields"], list))
check("build fields not empty",                 len(data["fields"]) > 0)

hr_field = next((f for f in data["fields"] if f["field"] == "heart_rate_series"), None)
check("heart_rate_series in fields",            hr_field is not None)
check("heart_rate series is list",              isinstance(hr_field["series"], list))
check("heart_rate series has entries",          len(hr_field["series"]) > 0)
check("heart_rate entry has ts",                "ts" in hr_field["series"][0])
check("heart_rate entry has value",             "value" in hr_field["series"][0])


# ══════════════════════════════════════════════════════════════════════════════
#  6. dash_plotter_html — render()
# ══════════════════════════════════════════════════════════════════════════════

section("6. dash_plotter_html — render()")

import dash_plotter_html as plotter_html

_out = _TMPDIR / "test_output.html"
plotter_html.render(data, _out, {})

check("output file written",                    _out.exists())
check("output file not empty",                  _out.stat().st_size > 0)

_html = _out.read_text(encoding="utf-8")
check("html contains DOCTYPE",                  "<!DOCTYPE html>" in _html)
check("html contains plotly script",            "plotly" in _html.lower())
check("html contains tab button",               "tab-btn" in _html)
check("html contains heart_rate chart",         "heart_rate_series" in _html)
check("html contains disclaimer",               "medical advice" in _html)
check("html contains footer",                   "github.com" in _html.lower())
check("html contains title",                    data["title"] in _html)

# render with empty fields raises ValueError
try:
    plotter_html.render({"title": "x", "subtitle": "", "fields": []}, _TMPDIR / "empty.html", {})
    check("render empty fields raises ValueError", False)
except ValueError:
    check("render empty fields raises ValueError", True)


# ══════════════════════════════════════════════════════════════════════════════
#  7. dash_runner — scan() and build()
# ══════════════════════════════════════════════════════════════════════════════

section("7. dash_runner — scan() and build()")

sys.path.insert(0, str(_ROOT / "dashboards"))
import dash_runner

specialists = dash_runner.scan()
check("scan finds at least one specialist",     len(specialists) > 0)

ts_spec = next((s for s in specialists if "Timeseries" in s["name"]), None)
check("timeseries specialist found",            ts_spec is not None)
check("timeseries has html format",             "html" in ts_spec["formats"])
check("timeseries has description",             ts_spec["description"] != "")

_out_dir = _TMPDIR / "dashboards_out"
_out_dir.mkdir()

selections = [(ts_spec["module"], "html")]
results = dash_runner.build(
    selections=selections,
    date_from=_TEST_DATE,
    date_to=_TEST_DATE,
    settings={},
    output_dir=_out_dir,
)
check("build returns results list",             isinstance(results, list))
check("build result count matches selections",  len(results) == 1)
check("build result success",                   results[0]["success"] is True)
check("build result file exists",               results[0]["file"].exists())

# Fehlerfall: ungültiges Format → success=False, kein Absturz
results_err = dash_runner.build(
    selections=[(ts_spec["module"], "nonexistent_format")],
    date_from=_TEST_DATE,
    date_to=_TEST_DATE,
    settings={},
    output_dir=_out_dir,
)
check("build invalid format: returns list",     isinstance(results_err, list))
check("build invalid format: success=False",    results_err[0]["success"] is False)
check("build invalid format: error key present", "error" in results_err[0])


# ══════════════════════════════════════════════════════════════════════════════
#  8. dash_plotter_excel — render()
# ══════════════════════════════════════════════════════════════════════════════

section("8. dash_plotter_excel — render()")

import dash_plotter_excel as plotter_excel
from openpyxl import load_workbook

_out_xlsx = _TMPDIR / "test_output.xlsx"
plotter_excel.render(data, _out_xlsx, {})

check("xlsx output file written",               _out_xlsx.exists())
check("xlsx output file not empty",             _out_xlsx.stat().st_size > 0)

_wb = load_workbook(_out_xlsx)
_sheet_names = _wb.sheetnames
check("xlsx has data sheet heart_rate",         "Heart Rate - Data"  in _sheet_names)
check("xlsx has chart sheet heart_rate",        "Heart Rate - Chart" in _sheet_names)
check("xlsx has data sheet stress",             "Stress - Data"      in _sheet_names)
check("xlsx has data sheet body_battery",       "Body Battery - Data" in _sheet_names)

_ws_hr = _wb["Heart Rate - Data"]
check("xlsx header row exists",                 _ws_hr.cell(1, 1).value == "Date")
check("xlsx data row has date",                 _ws_hr.cell(2, 1).value is not None)
check("xlsx data row has time",                 _ws_hr.cell(2, 2).value is not None)
check("xlsx data row has value",                _ws_hr.cell(2, 3).value == 58.0)

# render with empty fields raises ValueError
try:
    plotter_excel.render({"title": "x", "subtitle": "", "fields": []}, _TMPDIR / "empty.xlsx", {})
    check("xlsx render empty fields raises ValueError", False)
except ValueError:
    check("xlsx render empty fields raises ValueError", True)


# ══════════════════════════════════════════════════════════════════════════════
#  9. dash_plotter_json — render()
# ══════════════════════════════════════════════════════════════════════════════

section("9. dash_plotter_json — render()")

import dash_plotter_json as plotter_json

# Build minimal health-style data dict for JSON plotter
_health_data = {
    "title":           "Garmin Health Analysis",
    "date_from":       _TEST_DATE,
    "date_to":         _TEST_DATE,
    "prompt_template": "health_analysis",
    "profile": {
        "age": 35, "sex": "male", "vo2max": 48.0, "fitness": "excellent"
    },
    "fields": [
        {
            "field":         "hrv_last_night",
            "label":         "HRV",
            "unit":          "ms",
            "period_avg":    62.0,
            "baseline_avg":  58.3,
            "ref_low":       40.0,
            "ref_high":      85.0,
            "flagged_days":  2,
            "flagged_dates": ["2026-03-01", "2026-03-02"],
        },
        {
            "field":         "resting_heart_rate",
            "label":         "Resting HR",
            "unit":          "bpm",
            "period_avg":    55.0,
            "baseline_avg":  57.1,
            "ref_low":       40.0,
            "ref_high":      60.0,
            "flagged_days":  0,
            "flagged_dates": [],
        },
    ],
}

_out_json   = _TMPDIR / "test_health.json"
_out_prompt = _TMPDIR / "test_health_prompt.md"

plotter_json.render(_health_data, _out_json, {})

check("json file written",                      _out_json.exists())
check("prompt file written",                    _out_prompt.exists())
check("json file not empty",                    _out_json.stat().st_size > 0)
check("prompt file not empty",                  _out_prompt.stat().st_size > 0)

_jdata = json.loads(_out_json.read_text(encoding="utf-8"))
check("json has generated key",                 "generated" in _jdata)
check("json has profile",                       "profile" in _jdata)
check("json has fields",                        "fields" in _jdata)
check("json fields not empty",                  len(_jdata["fields"]) > 0)
check("json profile age correct",               _jdata["profile"]["age"] == 35)

_prompt_text = _out_prompt.read_text(encoding="utf-8")
check("prompt is markdown",                     "# Garmin Health Analysis" in _prompt_text)
check("prompt has profile section",             "## Personal Profile" in _prompt_text)
check("prompt has metric table",                "## Metric Summary" in _prompt_text)
check("prompt has flagged section",             "## Flagged Metrics" in _prompt_text)
check("prompt has instructions",                "## Instructions for Assistant" in _prompt_text)
check("prompt has disclaimer",                  "not medical advice" in _prompt_text)
check("prompt shows HRV flagged",               "HRV" in _prompt_text and "2" in _prompt_text)

# No template key — fallback prompt
_no_template = {**_health_data, "prompt_template": None}
_out_no_tpl  = _TMPDIR / "test_no_template.json"
plotter_json.render(_no_template, _out_no_tpl, {})
_fallback_prompt = (_TMPDIR / "test_no_template_prompt.md").read_text(encoding="utf-8")
check("fallback prompt written",                "No prompt template" in _fallback_prompt)

# Unknown template key — graceful fallback
_unknown_tpl = {**_health_data, "prompt_template": "nonexistent"}
_out_unk     = _TMPDIR / "test_unknown_template.json"
plotter_json.render(_unknown_tpl, _out_unk, {})
_unk_prompt  = (_TMPDIR / "test_unknown_template_prompt.md").read_text(encoding="utf-8")
check("unknown template: graceful fallback",    "No prompt template" in _unk_prompt)

# Template-Registry
import dash_prompt_templates as prompt_tmpl
check("templates list_templates not empty",     len(prompt_tmpl.list_templates()) > 0)
check("templates health_analysis registered",   "health_analysis" in prompt_tmpl.list_templates())
check("templates get returns callable",         callable(prompt_tmpl.get("health_analysis")))
try:
    prompt_tmpl.get("nonexistent")
    check("templates unknown raises KeyError",  False)
except KeyError:
    check("templates unknown raises KeyError",  True)


# ══════════════════════════════════════════════════════════════════════════════
#  10. health_garmin specialist — build()
# ══════════════════════════════════════════════════════════════════════════════

section("10. health_garmin specialist — build()")

# Synthetic summary data
_SUMMARY = {
    "date": _TEST_DATE,
    "sleep": {
        "hrv_last_night_ms": 62.0,
        "spo2_avg": 97.0,
        "duration_h": 7.5,
    },
    "heartrate": {"resting_bpm": 52.0},
    "stress":    {"stress_avg": 28.0, "body_battery_max": 85.0},
    "training":  {"vo2max": 48.0},
}

_sum_dir = _TMPDIR / "garmin_data" / "summary"
_sum_dir.mkdir(parents=True, exist_ok=True)
(_sum_dir / f"garmin_{_TEST_DATE}.json").write_text(
    json.dumps(_SUMMARY), encoding="utf-8"
)
# Vortag für Baseline-Berechnung
import datetime as _dt
_prev_date = (_dt.date.fromisoformat(_TEST_DATE) - _dt.timedelta(days=1)).isoformat()
(_sum_dir / f"garmin_{_prev_date}.json").write_text(
    json.dumps({**_SUMMARY, "date": _prev_date}), encoding="utf-8"
)
importlib.reload(cfg)

_health_spec = importlib.util.spec_from_file_location(
    "health_dash",
    _ROOT / "dashboards" / "health_garmin_html-json_dash.py"
)
_health_mod = importlib.util.module_from_spec(_health_spec)
_health_spec.loader.exec_module(_health_mod)

check("health META present",                    hasattr(_health_mod, "META"))
check("health META has html format",            "html" in _health_mod.META["formats"])
check("health META has json format",            "json" in _health_mod.META["formats"])

_settings = {"age": 35, "sex": "male"}
_hdata = _health_mod.build(_TEST_DATE, _TEST_DATE, _settings)

check("health build returns dict",              isinstance(_hdata, dict))
check("health build has profile",               "profile" in _hdata)
check("health build has fields",                "fields" in _hdata)
check("health build has date_from",             _hdata.get("date_from") == _TEST_DATE)
check("health build has prompt_template",       _hdata.get("prompt_template") == "health_analysis")
check("health profile has fitness",             "fitness" in _hdata["profile"])
check("health profile vo2max detected",         _hdata["profile"]["vo2max"] == 48.0)

_hrv = next((f for f in _hdata["fields"] if f["field"] == "hrv_last_night"), None)
check("hrv field present",                      _hrv is not None)
check("hrv has days",                           "days" in _hrv)
check("hrv has ref_low",                        "ref_low" in _hrv)
check("hrv has ref_high",                       "ref_high" in _hrv)
check("hrv has period_avg",                     _hrv.get("period_avg") == 62.0)
check("hrv day has date",                       _hrv["days"][0]["date"] == _TEST_DATE)
check("hrv day has value",                      _hrv["days"][0]["value"] == 62.0)
check("hrv day has baseline",                   "baseline" in _hrv["days"][0])
check("hrv day has status",                     "status" in _hrv["days"][0])

# HTML render with analysis data
_out_health_html = _TMPDIR / "test_health.html"
plotter_html.render(_hdata, _out_health_html, {})
_health_html = _out_health_html.read_text(encoding="utf-8")
check("health html written",                    _out_health_html.exists())
check("health html has hrv chart",              "hrv_last_night" in _health_html)
check("health html has reference band",         "tonexty" in _health_html)
check("health html has baseline trace",         "90d baseline" in _health_html)

# JSON render with analysis data
_out_health_json = _TMPDIR / "test_health2.json"
plotter_json.render(_hdata, _out_health_json, {})
_out_health_prompt = _TMPDIR / "test_health2_prompt.md"
check("health json written",                    _out_health_json.exists())
check("health prompt written",                  _out_health_prompt.exists())
_health_prompt = _out_health_prompt.read_text(encoding="utf-8")
check("health prompt has profile section",      "## Personal Profile" in _health_prompt)
check("health prompt has metric table",         "## Metric Summary" in _health_prompt)
check("health prompt has instructions",         "## Instructions" in _health_prompt)

# Summary ohne hrv-Feld — build() darf nicht abstürzen
_summary_no_hrv = {
    "date": _TEST_DATE,
    "sleep":    {"spo2_avg": 97.0, "duration_h": 7.5},
    "heartrate": {"resting_bpm": 52.0},
    "stress":    {"stress_avg": 28.0, "body_battery_max": 85.0},
    "training":  {"vo2max": 48.0},
}
(_sum_dir / f"garmin_{_TEST_DATE}.json").write_text(
    json.dumps(_summary_no_hrv), encoding="utf-8"
)
importlib.reload(cfg)
_hdata_no_hrv = _health_mod.build(_TEST_DATE, _TEST_DATE, _settings)
check("health build no hrv: returns dict",      isinstance(_hdata_no_hrv, dict))
check("health build no hrv: has fields",        "fields" in _hdata_no_hrv)
_hrv_absent = next((f for f in _hdata_no_hrv["fields"] if f["field"] == "hrv_last_night"), None)
check("health build no hrv: field absent or value None",
      _hrv_absent is None or _hrv_absent.get("days", [{}])[0].get("value") is None)

# Isolation: Original-Summary wiederherstellen für folgende Sections
(_sum_dir / f"garmin_{_TEST_DATE}.json").write_text(
    json.dumps(_SUMMARY), encoding="utf-8"
)

# Flag guard: sleep_duration = 0.0 → val must be None in days
_summary_zero_sleep = {**_SUMMARY, "sleep": {**_SUMMARY["sleep"], "duration_h": 0.0}}
(_sum_dir / f"garmin_{_TEST_DATE}.json").write_text(
    json.dumps(_summary_zero_sleep), encoding="utf-8"
)
importlib.reload(cfg)
_hdata_zero_sleep = _health_mod.build(_TEST_DATE, _TEST_DATE, _settings)
_sleep_field = next((f for f in _hdata_zero_sleep["fields"] if f["field"] == "sleep_duration"), None)
check("flag guard: sleep_duration 0.0 → val is None",
      _sleep_field is not None and _sleep_field["days"][0]["value"] is None)

# Auto-size: date_from before any data → d_from adjusted, subtitle contains 'adjusted'
_early_date = "2000-01-01"
importlib.reload(cfg)
_hdata_autosize = _health_mod.build(_early_date, _TEST_DATE, _settings)
check("auto-size: subtitle present",               "subtitle" in _hdata_autosize)
check("auto-size: subtitle shows adjusted range",  "adjusted" in _hdata_autosize.get("subtitle", ""))

# Isolation: Original-Summary wiederherstellen
(_sum_dir / f"garmin_{_TEST_DATE}.json").write_text(
    json.dumps(_SUMMARY), encoding="utf-8"
)
importlib.reload(cfg)


# ══════════════════════════════════════════════════════════════════════════════
#  11. overview_garmin specialist — build() + excel render
# ══════════════════════════════════════════════════════════════════════════════

section("11. overview_garmin specialist — build()")

_overview_spec = importlib.util.spec_from_file_location(
    "overview_dash",
    _ROOT / "dashboards" / "overview_garmin_xls_dash.py"
)
_overview_mod = importlib.util.module_from_spec(_overview_spec)
_overview_spec.loader.exec_module(_overview_mod)

check("overview META present",                  hasattr(_overview_mod, "META"))
check("overview META has excel format",         "excel" in _overview_mod.META["formats"])

_odata = _overview_mod.build(_TEST_DATE, _TEST_DATE, {})
check("overview build returns dict",            isinstance(_odata, dict))
check("overview build has rows",                "rows" in _odata)
check("overview build has columns",             "columns" in _odata)
check("overview rows not empty",                len(_odata["rows"]) > 0)
check("overview row has date",                  _odata["rows"][0]["date"] == _TEST_DATE)
check("overview row has values",                "values" in _odata["rows"][0])
check("overview columns has sleep_duration",    any(c["field"] == "sleep_duration" for c in _odata["columns"]))

# Multi-Tage: zweiter Tag schon vorhanden via _prev_date aus Sek. 10
_odata_multi = _overview_mod.build(_prev_date, _TEST_DATE, {})
check("overview multi: 2 rows",                 len(_odata_multi["rows"]) == 2)
check("overview multi: rows sorted asc",        _odata_multi["rows"][0]["date"] <= _odata_multi["rows"][1]["date"])

_out_overview = _TMPDIR / "test_overview.xlsx"
plotter_excel.render(_odata, _out_overview, {})
check("overview xlsx written",                  _out_overview.exists())
check("overview xlsx not empty",                _out_overview.stat().st_size > 0)

_wb_ov = load_workbook(_out_overview)
check("overview has sheet",                     len(_wb_ov.sheetnames) > 0)
check("overview sheet has header row",          _wb_ov.active.cell(1, 1).value == "Date")
check("overview sheet has data row",            _wb_ov.active.cell(2, 1).value == _TEST_DATE)

# Excel render with analysis data (days mode)
_out_health_xlsx = _TMPDIR / "test_health.xlsx"
plotter_excel.render(_hdata, _out_health_xlsx, {})
_wb_health = load_workbook(_out_health_xlsx)
check("health xlsx written",                    _out_health_xlsx.exists())
check("health xlsx has hrv sheet",              any("HRV" in s for s in _wb_health.sheetnames))

# render empty overview raises ValueError
try:
    plotter_excel.render({"rows": []}, _TMPDIR / "empty_ov.xlsx", {})
    check("overview empty rows raises ValueError", False)
except ValueError:
    check("overview empty rows raises ValueError", True)


# ══════════════════════════════════════════════════════════════════════════════
#  12. health_garmin-weather-pollen specialist — build()
# ══════════════════════════════════════════════════════════════════════════════

section("12. health_garmin-weather-pollen specialist — build()")

# Synthetic context data
_ctx_weather_dir = _TMPDIR / "context_data" / "weather" / "raw"
_ctx_pollen_dir  = _TMPDIR / "context_data" / "pollen"  / "raw"
_ctx_weather_dir.mkdir(parents=True, exist_ok=True)
_ctx_pollen_dir.mkdir(parents=True, exist_ok=True)

(_ctx_weather_dir / f"weather_{_TEST_DATE}.json").write_text(
    json.dumps({"date": _TEST_DATE, "fields": {
        "temperature_2m_max": 18.5,
        "uv_index_max": 4.2,
    }}), encoding="utf-8"
)
(_ctx_pollen_dir / f"pollen_{_TEST_DATE}.json").write_text(
    json.dumps({"date": _TEST_DATE, "fields": {
        "birch_pollen_index": 45.2,
    }}), encoding="utf-8"
)
importlib.reload(cfg)

_ctx_spec = importlib.util.spec_from_file_location(
    "health_context_dash",
    _ROOT / "dashboards" / "health_garmin-weather-pollen_html-xls_dash.py"
)
_ctx_mod = importlib.util.module_from_spec(_ctx_spec)
_ctx_spec.loader.exec_module(_ctx_mod)

check("context META present",                   hasattr(_ctx_mod, "META"))
check("context META has html format",           "html" in _ctx_mod.META["formats"])
check("context META has excel format",          "excel" in _ctx_mod.META["formats"])

_cdata = _ctx_mod.build(_TEST_DATE, _TEST_DATE, {})
check("context build returns dict",             isinstance(_cdata, dict))
check("context build has fields",               "fields" in _cdata)
check("context fields not empty",               len(_cdata["fields"]) > 0)

_hrv_f   = next((f for f in _cdata["fields"] if f["field"] == "hrv_last_night"), None)
_temp_f  = next((f for f in _cdata["fields"] if f["field"] == "temperature_max"), None)
_pollen_f= next((f for f in _cdata["fields"] if f["field"] == "pollen_birch"), None)

check("garmin field hrv present",               _hrv_f is not None)
check("context field temperature present",      _temp_f is not None)
check("context field pollen present",           _pollen_f is not None)
check("hrv has days",                           "days" in _hrv_f)
check("hrv group is garmin",                    _hrv_f.get("group") == "garmin")
check("temperature group is weather",           _temp_f.get("group") == "weather")
check("pollen group is pollen",                 _pollen_f.get("group") == "pollen")
check("hrv day has date",                       _hrv_f["days"][0]["date"] == _TEST_DATE)
check("hrv day value correct",                  _hrv_f["days"][0]["value"] == 62.0)

# HTML render
_out_ctx_html = _TMPDIR / "test_context.html"
plotter_html.render(_cdata, _out_ctx_html, {})
_ctx_html = _out_ctx_html.read_text(encoding="utf-8")
check("context html written",                   _out_ctx_html.exists())
check("context html has hrv chart",             "hrv_last_night" in _ctx_html)
check("context html has temperature chart",     "temperature_max" in _ctx_html)

# Excel render
_out_ctx_xlsx = _TMPDIR / "test_context.xlsx"
plotter_excel.render(_cdata, _out_ctx_xlsx, {})
check("context xlsx written",                   _out_ctx_xlsx.exists())
_wb_ctx = load_workbook(_out_ctx_xlsx)
check("context xlsx has hrv sheet",             any("HRV" in s for s in _wb_ctx.sheetnames))
check("context xlsx has temp sheet",            any("Temp" in s for s in _wb_ctx.sheetnames))


# ══════════════════════════════════════════════════════════════════════════════
#  13. sleep_recovery_context specialist — build() + complex plotter
# ══════════════════════════════════════════════════════════════════════════════

section("13. sleep_recovery_context specialist — build() + complex plotter")

import dash_plotter_html_complex as plotter_complex

_src_spec = importlib.util.spec_from_file_location(
    "sleep_recovery_dash",
    _ROOT / "dashboards" / "sleep_recovery_context_dash.py"
)
_src_mod = importlib.util.module_from_spec(_src_spec)
_src_spec.loader.exec_module(_src_mod)

check("sleep_recovery META present",             hasattr(_src_mod, "META"))
check("sleep_recovery META has html_complex",    "html_complex" in _src_mod.META["formats"])

_srdata = _src_mod.build(_TEST_DATE, _TEST_DATE, {})
check("sleep_recovery build returns dict",       isinstance(_srdata, dict))
check("sleep_recovery has daily key",            "daily" in _srdata)
check("sleep_recovery has intraday key",         "intraday" in _srdata)

_daily = _srdata["daily"]
check("daily has dates",                         "dates" in _daily and len(_daily["dates"]) > 0)
check("daily has hrv list",                      "hrv" in _daily)
check("daily has body_battery list",             "body_battery" in _daily)
check("daily has sleep_h list",                  "sleep_h" in _daily)
check("daily has sleep_phases",                  "sleep_phases" in _daily)

_phase = _daily["sleep_phases"][0]
check("sleep_phases entry has date",             "date" in _phase)
check("sleep_phases has deep",                   "deep" in _phase)
check("sleep_phases has rem",                    "rem" in _phase)

# raw_pct calculation: deepSleepSeconds=5400, sleepTimeSeconds=27000 → 20.0%
check("deep_pct calculated correctly",           _phase.get("deep") == 20.0)
# lightSleepSeconds=13500 / 27000 → 50.0%
check("light_pct calculated correctly",          _phase.get("light") == 50.0)
# remSleepSeconds=6750 / 27000 → 25.0%
check("rem_pct calculated correctly",            _phase.get("rem") == 25.0)
# awakeSleepSeconds=1350 / 27000 → 5.0%
check("awake_pct calculated correctly",          _phase.get("awake") == 5.0)

_intraday = _srdata["intraday"]
check("intraday has test date",                  _TEST_DATE in _intraday)
check("intraday day has heart_rate key",         "heart_rate" in _intraday[_TEST_DATE])
check("intraday day has stress key",             "stress" in _intraday[_TEST_DATE])
check("intraday day has body_battery key",       "body_battery" in _intraday[_TEST_DATE])

# Complex plotter render
_out_src_html = _TMPDIR / "test_sleep_recovery.html"
plotter_complex.render(_srdata, _out_src_html, {})
_src_html = _out_src_html.read_text(encoding="utf-8")
check("sleep_recovery html written",             _out_src_html.exists())
check("sleep_recovery html not empty",           _out_src_html.stat().st_size > 0)
check("sleep_recovery html has DOCTYPE",         "<!DOCTYPE html>" in _src_html)
check("sleep_recovery html has tab1",            "chart-tab1" in _src_html)
check("sleep_recovery html has tab2",            "chart-tab2" in _src_html)
check("sleep_recovery html has plotly",          "plotly" in _src_html.lower())
check("sleep_recovery html has disclaimer",      "medical advice" in _src_html)
check("sleep_recovery html has deep sleep",      "Deep Sleep" in _src_html)
check("sleep_recovery html has intraday func",   "updateIntradayChart" in _src_html)

# render without daily/intraday raises ValueError
try:
    plotter_complex.render({"title": "x"}, _TMPDIR / "empty_src.html", {})
    check("complex render missing keys raises ValueError", False)
except ValueError:
    check("complex render missing keys raises ValueError", True)


# ══════════════════════════════════════════════════════════════════════════════
#  Cleanup + summary
# ══════════════════════════════════════════════════════════════════════════════

shutil.rmtree(_TMPDIR, ignore_errors=True)

total = _pass + _fail
print(f"\n{'═' * 55}")
print(f"  {total} checks — {_pass} passed, {_fail} failed")
if _failures:
    print(f"\n  Failed:")
    for f in _failures:
        print(f"    ✗  {f}")
print(f"{'═' * 55}\n")

sys.exit(0 if _fail == 0 else 1)