#!/usr/bin/env python3
"""
garmin_to_excel.py

Reads garmin_YYYY-MM-DD.json files from the summary folder and exports
selected fields as a formatted Excel spreadsheet.

Configuration: only edit the CONFIG block below, leave the rest untouched.
"""

import json
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
#  CONFIG  –  edit here, do not touch anything below
# ══════════════════════════════════════════════════════════════════════════════

# Path to the summary folder (same as in garmin_collector.py)
SUMMARY_DIR = Path("~/garmin_data/summary").expanduser()

# Output file
OUTPUT_FILE = Path("~/garmin_export.xlsx").expanduser()

# Date range (None = export everything available)
DATE_FROM = None   # e.g. "2024-01-01"
DATE_TO   = None   # e.g. "2024-12-31"

# ── Toggle columns on/off ──────────────────────────────────────────────────────
# True  = column appears in the spreadsheet
# False = column is omitted
FIELDS = {
    # ── Sleep ────────────────────────────────────────
    "sleep.duration_h":          True,   # Total sleep duration (h)
    "sleep.deep_h":              True,   # Deep sleep (h)
    "sleep.rem_h":               True,   # REM sleep (h)
    "sleep.light_h":             False,  # Light sleep (h)
    "sleep.awake_h":             False,  # Awake time during night (h)
    "sleep.score":               True,   # Sleep score
    "sleep.spo2_avg":            True,   # Avg SpO2 (%)
    "sleep.respiration_avg":     False,  # Avg respiration rate
    "sleep.hrv_last_night_ms":   True,   # HRV last night (ms)
    "sleep.hrv_weekly_avg_ms":   True,   # HRV weekly average (ms)
    "sleep.hrv_status":          False,  # HRV status (text)
    "sleep.hrv_feedback":        False,  # HRV feedback (long text)

    # ── Heart rate ───────────────────────────────────
    "heartrate.resting_bpm":     True,   # Resting heart rate
    "heartrate.max_bpm":         True,   # Daily maximum
    "heartrate.min_bpm":         False,  # Daily minimum
    "heartrate.avg_bpm":         False,  # Daily average

    # ── Stress & Body Battery ────────────────────────
    "stress.stress_avg":         True,   # Avg stress level
    "stress.stress_max":         False,  # Max stress
    "stress.body_battery_max":   True,   # Body Battery maximum
    "stress.body_battery_min":   False,  # Body Battery minimum
    "stress.body_battery_end":   True,   # Body Battery end of day

    # ── Daily stats ──────────────────────────────────
    "day.steps":                 True,   # Steps
    "day.steps_goal":            False,  # Step goal
    "day.calories_active":       True,   # Active calories (kcal)
    "day.calories_total":        False,  # Total calories (kcal)
    "day.intensity_min_moderate":True,   # Moderate intensity minutes
    "day.intensity_min_vigorous":True,   # Vigorous intensity minutes
    "day.floors_climbed":        False,  # Floors climbed
    "day.distance_km":           True,   # Distance covered (km)

    # ── Training ─────────────────────────────────────
    "training.readiness_score":  True,   # Training readiness score
    "training.readiness_level":  False,  # Training readiness level (text)
    "training.readiness_feedback":False, # Readiness feedback (long text)
    "training.training_status":  False,  # Training status (text)
    "training.training_load_7d": True,   # 7-day training load
    "training.vo2max":           True,   # VO2max
}

# Export activities as a separate sheet?
EXPORT_ACTIVITIES_SHEET = True

# ══════════════════════════════════════════════════════════════════════════════
#  Column labels – change if desired
# ══════════════════════════════════════════════════════════════════════════════

LABELS = {
    "sleep.duration_h":           "Sleep total (h)",
    "sleep.deep_h":               "Deep sleep (h)",
    "sleep.rem_h":                "REM (h)",
    "sleep.light_h":              "Light sleep (h)",
    "sleep.awake_h":              "Awake (h)",
    "sleep.score":                "Sleep score",
    "sleep.spo2_avg":             "SpO2 avg (%)",
    "sleep.respiration_avg":      "Respiration avg",
    "sleep.hrv_last_night_ms":    "HRV night (ms)",
    "sleep.hrv_weekly_avg_ms":    "HRV week avg (ms)",
    "sleep.hrv_status":           "HRV status",
    "sleep.hrv_feedback":         "HRV feedback",
    "heartrate.resting_bpm":      "Resting HR (bpm)",
    "heartrate.max_bpm":          "HR max (bpm)",
    "heartrate.min_bpm":          "HR min (bpm)",
    "heartrate.avg_bpm":          "HR avg (bpm)",
    "stress.stress_avg":          "Stress avg",
    "stress.stress_max":          "Stress max",
    "stress.body_battery_max":    "Body Battery max",
    "stress.body_battery_min":    "Body Battery min",
    "stress.body_battery_end":    "Body Battery EOD",
    "day.steps":                  "Steps",
    "day.steps_goal":             "Steps goal",
    "day.calories_active":        "Calories active",
    "day.calories_total":         "Calories total",
    "day.intensity_min_moderate": "Int. min moderate",
    "day.intensity_min_vigorous": "Int. min vigorous",
    "day.floors_climbed":         "Floors",
    "day.distance_km":            "Distance (km)",
    "training.readiness_score":   "Readiness score",
    "training.readiness_level":   "Readiness level",
    "training.readiness_feedback":"Readiness feedback",
    "training.training_status":   "Training status",
    "training.training_load_7d":  "Training load 7d",
    "training.vo2max":            "VO2max",
}

# ══════════════════════════════════════════════════════════════════════════════
#  Styles
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", start_color="1F3864")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATE_FONT    = Font(name="Arial", bold=True, size=10)
DATA_FONT    = Font(name="Arial", size=10)
BORDER_SIDE  = Side(style="thin", color="D0D0D0")
CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                      top=BORDER_SIDE,  bottom=BORDER_SIDE)

GROUP_COLORS = {
    "sleep":     "DDEEFF",
    "heartrate": "FFE8CC",
    "stress":    "E8FFE8",
    "day":       "FFF8CC",
    "training":  "F0E8FF",
}


def group_fill(key: str) -> PatternFill:
    group = key.split(".")[0]
    color = GROUP_COLORS.get(group, "FFFFFF")
    return PatternFill("solid", start_color=color)


# ══════════════════════════════════════════════════════════════════════════════
#  Data loading
# ══════════════════════════════════════════════════════════════════════════════

def load_summaries() -> list[dict]:
    date_from = date.fromisoformat(DATE_FROM) if DATE_FROM else None
    date_to   = date.fromisoformat(DATE_TO)   if DATE_TO   else None

    rows = []
    for f in sorted(SUMMARY_DIR.glob("garmin_*.json")):
        try:
            d = date.fromisoformat(f.stem.replace("garmin_", ""))
        except ValueError:
            continue
        if date_from and d < date_from:
            continue
        if date_to   and d > date_to:
            continue
        with open(f, encoding="utf-8") as fp:
            rows.append(json.load(fp))

    print(f"  {len(rows)} days loaded from {SUMMARY_DIR}")
    return rows


def get_val(summary: dict, dotkey: str):
    section, field = dotkey.split(".", 1)
    return (summary.get(section) or {}).get(field)


# ══════════════════════════════════════════════════════════════════════════════
#  Excel builder
# ══════════════════════════════════════════════════════════════════════════════

def build_excel(summaries: list[dict]):
    wb = Workbook()

    # ── Main sheet ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Garmin Daily Overview"
    ws.freeze_panes = "B2"

    active_fields = [k for k, v in FIELDS.items() if v]

    ws.cell(1, 1, "Date").font      = HEADER_FONT
    ws.cell(1, 1).fill              = HEADER_FILL
    ws.cell(1, 1).alignment         = Alignment(horizontal="center")
    ws.cell(1, 1).border            = CELL_BORDER
    ws.column_dimensions["A"].width = 12

    for col_idx, key in enumerate(active_fields, start=2):
        cell = ws.cell(1, col_idx, LABELS.get(key, key))
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = CELL_BORDER

    ws.row_dimensions[1].height = 36

    for row_idx, s in enumerate(summaries, start=2):
        date_cell = ws.cell(row_idx, 1, s.get("date"))
        date_cell.font      = DATE_FONT
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.border    = CELL_BORDER

        for col_idx, key in enumerate(active_fields, start=2):
            val  = get_val(s, key)
            cell = ws.cell(row_idx, col_idx, val)
            cell.font      = DATA_FONT
            cell.fill      = group_fill(key)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = CELL_BORDER

            if isinstance(val, float):
                cell.number_format = "0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"

        if row_idx % 2 == 0:
            for col_idx in range(1, len(active_fields) + 2):
                c = ws.cell(row_idx, col_idx)
                if c.fill.fgColor.rgb == "00000000":
                    c.fill = PatternFill("solid", start_color="F5F5F5")

    for col_idx, key in enumerate(active_fields, start=2):
        label_len = len(LABELS.get(key, key))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(label_len + 2, 20))

    # ── Activities sheet ─────────────────────────────────────────────────────
    if EXPORT_ACTIVITIES_SHEET:
        wa = wb.create_sheet("Activities")
        wa.freeze_panes = "A2"

        act_headers = ["Date", "Name", "Type", "Duration (min)", "Distance (km)",
                       "HR avg", "HR max", "Calories", "TE aerobic", "TE anaerobic"]
        act_widths  = [12, 24, 18, 14, 13, 8, 8, 10, 11, 12]

        for col_idx, (h, w) in enumerate(zip(act_headers, act_widths), start=1):
            cell = wa.cell(1, col_idx, h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border    = CELL_BORDER
            wa.column_dimensions[get_column_letter(col_idx)].width = w

        wa.row_dimensions[1].height = 30
        row_idx  = 2
        ACT_FILL = PatternFill("solid", start_color="F0E8FF")

        for s in summaries:
            for a in (s.get("activities") or []):
                vals = [
                    s.get("date"),
                    a.get("name"),
                    a.get("type"),
                    a.get("duration_min"),
                    a.get("distance_km"),
                    a.get("avg_hr"),
                    a.get("max_hr"),
                    a.get("calories"),
                    a.get("training_effect_aerobic"),
                    a.get("training_effect_anaerobic"),
                ]
                for col_idx, v in enumerate(vals, start=1):
                    cell = wa.cell(row_idx, col_idx, v)
                    cell.font      = DATA_FONT
                    cell.fill      = ACT_FILL
                    cell.alignment = Alignment(horizontal="center")
                    cell.border    = CELL_BORDER
                    if isinstance(v, float):
                        cell.number_format = "0.00"
                row_idx += 1

    return wb


# ══════════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("Garmin → Excel export")
    print(f"  Source:  {SUMMARY_DIR}")
    print(f"  Output:  {OUTPUT_FILE}")

    if not SUMMARY_DIR.exists():
        print(f"  ERROR: folder not found: {SUMMARY_DIR}")
        return

    summaries = load_summaries()
    if not summaries:
        print("  No data found.")
        return

    print(f"  Active columns: {sum(v for v in FIELDS.values())}")
    wb = build_excel(summaries)

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"  ✓ Saved: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
